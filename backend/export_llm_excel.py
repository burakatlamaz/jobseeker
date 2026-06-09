from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import os
import re

from config import FINAL_EXCEL_FILE, FILTERED_DIR, LLM_RESULTS_JSONL_FILE
from utils import ensure_directories, read_jsonl_file


def safe(value) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(item) for item in value if item)
    return str(value).strip()


def best_link(row: dict) -> str:
    return row.get("apply_url") or row.get("source_job_url") or row.get("url") or ""


ROLE_PATTERNS = [
    ("internship", re.compile(r"\b(praktikum|praktikant|internship|intern)\b", re.IGNORECASE)),
    ("working_student", re.compile(r"\b(werkstudent|werksstudent|working student|werkstudium)\b", re.IGNORECASE)),
    ("hiwi", re.compile(r"\b(hiwi|studentische hilfskraft|student assistant|research assistant)\b", re.IGNORECASE)),
    ("thesis", re.compile(r"\b(masterarbeit|bachelorarbeit|thesis)\b", re.IGNORECASE)),
]


def role_type(row: dict) -> str:
    text = " ".join(
        safe(row.get(key))
        for key in [
            "title",
            "employment_type_raw",
            "llm_reason",
            "description",
            "description_raw",
        ]
    )
    for label, pattern in ROLE_PATTERNS:
        if pattern.search(text):
            return label

    employment = safe(row.get("employment_type_raw")).casefold()
    if "full" in employment or "vollzeit" in employment:
        return "full_time_or_regular"
    if "part" in employment or "teilzeit" in employment:
        return "part_time_regular"
    return "other"


def autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 42)


def write_workbook(path, rows: list[dict]) -> None:
    rows.sort(key=lambda row: int(row.get("llm_score") or 0), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Apply"

    headers = [
        "Status",
        "Score",
        "Category",
        "Role Type",
        "Title",
        "Company",
        "Location",
        "Role Source",
        "Apply Link",
        "Reason",
        "Tags",
        "Notes",
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for row in rows:
        ws.append(
            [
                "todo",
                int(row.get("llm_score") or 0),
                safe(row.get("llm_category")),
                role_type(row),
                safe(row.get("title")),
                safe(row.get("company")),
                safe(row.get("location_raw")),
                safe(row.get("sources")),
                best_link(row),
                safe(row.get("llm_reason")),
                safe(row.get("llm_fit_tags")),
                "",
            ]
        )

    for row_index in range(2, ws.max_row + 1):
        link_cell = ws.cell(row_index, 9)
        if link_cell.value:
            link_cell.hyperlink = link_cell.value
            link_cell.style = "Hyperlink"
        for cell in ws[row_index]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

    print(f"[INFO] Excel written to: {path}")
    print(f"[INFO] Rows written: {len(rows)}")


def main() -> None:
    ensure_directories([FILTERED_DIR])
    run_label = safe(os.environ.get("JOB_RUN_LABEL")).replace(" ", "_").lower()
    rows = [
        row
        for row in read_jsonl_file(LLM_RESULTS_JSONL_FILE)
        if row.get("llm_apply") is True and not row.get("llm_error")
    ]

    final_excel_file = (
        FILTERED_DIR / f"job_applications_{run_label}.xlsx"
        if run_label
        else FINAL_EXCEL_FILE
    )
    write_workbook(final_excel_file, rows)

    split_outputs = {
        "priority": [row for row in rows if int(row.get("llm_score") or 0) >= 7],
        "working_student": [row for row in rows if role_type(row) == "working_student"],
        "internship": [row for row in rows if role_type(row) == "internship"],
        "hiwi": [row for row in rows if role_type(row) == "hiwi"],
        "thesis": [row for row in rows if role_type(row) == "thesis"],
        "regular_or_other": [
            row
            for row in rows
            if role_type(row) in {"full_time_or_regular", "part_time_regular", "other"}
        ],
    }
    for name, split_rows in split_outputs.items():
        suffix = f"{run_label}_{name}" if run_label else name
        write_workbook(FILTERED_DIR / f"job_applications_{suffix}.xlsx", split_rows)


if __name__ == "__main__":
    main()
